from __future__ import annotations
from pathlib import Path
from typing import Any
import csv
import json
import re

try:
    import yaml  # type: ignore
except Exception:
    yaml = None

SUPPORTED = {'.md', '.txt', '.json', '.yaml', '.yml', '.csv', '.xlsx'}


def _safe_read_text(path: Path) -> str:
    return path.read_text(encoding='utf-8', errors='ignore')


def _extract_traits(text: str) -> list[str]:
    words = re.findall(r"[A-Za-z][A-Za-z0-9_-]{3,}", text.lower())
    stop = {'this', 'that', 'with', 'from', 'into', 'card', 'cards', 'agent', 'service', 'model'}
    out: list[str] = []
    for w in words:
        if w in stop:
            continue
        if w not in out:
            out.append(w)
        if len(out) >= 6:
            break
    return out


def _kv_from_text(text: str) -> dict[str, str]:
    kv = {}
    for line in text.splitlines():
        if ':' in line and len(line) < 220:
            k, v = line.split(':', 1)
            k = k.strip().lower().replace(' ', '_')
            v = v.strip()
            if k and v:
                kv[k] = v
    return kv


def _card_type_from_path(path: Path, data_keys: set[str] | None = None) -> str:
    p = str(path).lower()
    keys = data_keys or set()
    if 'service' in p or {'port', 'host', 'server', 'app_name'} & keys:
        return 'service'
    if 'agent' in p or {'model', 'personality', 'heartbeat', 'soul'} & keys:
        return 'agent'
    if path.suffix.lower() in {'.csv', '.xlsx'}:
        return 'dataset'
    if 'workspace' in p:
        return 'workspace'
    return 'generic'


def _dataset_section(path: Path, rows: int, headers: list[str]) -> dict[str, Any]:
    x = headers[0] if headers else 'index'
    ys = headers[1:2] if len(headers) > 1 else []
    return {
        'kind': 'chart',
        'title': f'{path.stem} preview',
        'content': {'chartType': 'line', 'x': x, 'y': ys}
    }


def parse_file_to_card(path: Path, root: Path) -> dict[str, Any] | None:
    ext = path.suffix.lower()
    rel = str(path.relative_to(root))
    if ext not in SUPPORTED:
        return None

    base = {
        'id': f'card-{abs(hash(str(path))) % 10**10}',
        'title': path.stem.replace('_', ' ').title(),
        'subtitle': rel,
        'description': f'Generated from {rel}',
        'stats': [],
        'traits': [],
        'sections': [],
        'source': {'path': str(path), 'parser': f'{ext[1:]}-parser', 'confidence': 0.75},
        'editableFields': ['title', 'subtitle', 'description', 'traits', 'sections'],
    }

    if ext in {'.txt', '.md'}:
        text = _safe_read_text(path)
        kv = _kv_from_text(text)
        ctype = _card_type_from_path(path, set(kv.keys()))
        base['type'] = ctype
        base['traits'] = _extract_traits(text)
        if kv:
            base['sections'].append({'kind': 'kv', 'title': 'Extracted Fields', 'content': kv})
            for k in ['model', 'personality', 'heartbeat', 'soul', 'port', 'host', 'server']:
                if k in kv:
                    base['stats'].append({'label': k.title(), 'value': kv[k]})
        else:
            snippet = "\n".join(text.splitlines()[:12]).strip()
            base['sections'].append({'kind': 'markdown', 'title': 'Preview', 'content': snippet})
        return base

    if ext == '.json':
        obj = json.loads(_safe_read_text(path))
        if isinstance(obj, dict):
            keys = set(obj.keys())
            base['type'] = _card_type_from_path(path, keys)
            base['sections'].append({'kind': 'kv', 'title': 'JSON Fields', 'content': obj})
            for k in ['model', 'personality', 'heartbeat', 'soul', 'port', 'host', 'server', 'app_name']:
                if k in obj:
                    base['stats'].append({'label': k.title(), 'value': obj[k]})
            base['traits'] = [k for k in list(keys)[:6]]
        else:
            base['type'] = 'generic'
            base['sections'].append({'kind': 'markdown', 'title': 'JSON Preview', 'content': str(obj)[:500]})
        return base

    if ext in {'.yaml', '.yml'}:
        if yaml is None:
            return None
        obj = yaml.safe_load(_safe_read_text(path))
        if not isinstance(obj, dict):
            obj = {'content': str(obj)}
        keys = set(obj.keys())
        base['type'] = _card_type_from_path(path, keys)
        base['sections'].append({'kind': 'kv', 'title': 'YAML Fields', 'content': obj})
        for k in ['model', 'personality', 'heartbeat', 'soul', 'port', 'host', 'server', 'app_name']:
            if k in obj:
                base['stats'].append({'label': k.title(), 'value': obj[k]})
        base['traits'] = [k for k in list(keys)[:6]]
        return base

    if ext == '.csv':
        with path.open('r', encoding='utf-8', errors='ignore', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)
        headers = rows[0] if rows else []
        body = rows[1:6] if len(rows) > 1 else []
        base['type'] = 'dataset'
        base['stats'] = [
            {'label': 'Rows', 'value': max(0, len(rows)-1)},
            {'label': 'Columns', 'value': len(headers)},
        ]
        base['sections'].append({'kind': 'table', 'title': 'Sample Rows', 'content': {'headers': headers, 'rows': body}})
        base['sections'].append(_dataset_section(path, max(0, len(rows)-1), headers))
        base['traits'] = ['dataset', 'csv']
        base['source']['confidence'] = 0.9
        return base

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return None
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else '' for h in (values[0] if values else [])]
        preview = [list(r) for r in values[1:6]] if len(values) > 1 else []
        base['type'] = 'dataset'
        base['stats'] = [
            {'label': 'Rows', 'value': max(0, len(values)-1)},
            {'label': 'Columns', 'value': len(headers)},
        ]
        base['sections'].append({'kind': 'table', 'title': 'Sample Rows', 'content': {'headers': headers, 'rows': preview}})
        base['sections'].append(_dataset_section(path, max(0, len(values)-1), headers))
        base['traits'] = ['dataset', 'xlsx']
        base['source']['confidence'] = 0.88
        return base

    return None


def ingest_folder(folder: str) -> dict[str, Any]:
    root = Path(folder).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        return {'status': 'error', 'message': 'folder not found', 'path': str(root)}

    cards = []
    skipped = []
    errors = []

    for path in root.rglob('*'):
        if not path.is_file():
            continue
        if path.suffix.lower() not in SUPPORTED:
            continue
        try:
            card = parse_file_to_card(path, root)
            if card is None:
                skipped.append(str(path))
            else:
                cards.append(card)
        except Exception as e:
            errors.append({'path': str(path), 'error': str(e)})

    return {
        'status': 'ok',
        'path': str(root),
        'cards': cards,
        'report': {
            'parsed': len(cards),
            'skipped': len(skipped),
            'errors': len(errors),
            'error_details': errors[:20],
        }
    }
